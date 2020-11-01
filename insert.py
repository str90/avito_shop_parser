import pymysql
from openpyxl import Workbook

startExportId = 919

mysqlCred = {
    'hostname': 'localhost',
    'username': 'phpmyadmin',
    'password': 'some_pass',
    'database': 'matrasspl'
}


def getLastDbId(mysqlCred):
    mysqlConnection = pymysql.connect(mysqlCred['hostname'], mysqlCred['username'],
                                      mysqlCred['password'], mysqlCred['database'])
    try:
        with mysqlConnection.cursor() as cursor:
            sql = cursor.execute('select LAST_INSERT_ID() FROM `avito_items`')
        mysqlConnection.commit()
    finally:
        mysqlConnection.close()
    if sql is not None:
        return int(sql)
    else:
        return 0


def getItemInfo(id, mysqlCred):
    mysqlConnection = pymysql.connect(mysqlCred['hostname'], mysqlCred['username'],
                                      mysqlCred['password'], mysqlCred['database'])
    try:
        with mysqlConnection.cursor() as cursor:
            sql = "SELECT * FROM `avito_items` WHERE `id`=%s"
            cursor.execute(sql, id)
            item_info_query = cursor.fetchone()

            sql = "SELECT * FROM `img_urls` WHERE `id`=%s"
            cursor.execute(sql, id)
            item_imgs_query = cursor.fetchall()
            imgs_list = list()
            for image in item_imgs_query:
                imgs_list.append(image[1])
    finally:
        mysqlConnection.close()
    returnDict = {
        'name': item_info_query[1],
        'description': item_info_query[2],
        'price_rub': item_info_query[3],
        'price_zl': item_info_query[4],
        'imgs_list': imgs_list
    }
    return returnDict


excelBookHandler = Workbook()
mainSheet = excelBookHandler.active
imageSheet = excelBookHandler.create_sheet("Images")
mainImageSheet = excelBookHandler.create_sheet("Main_images")
mainSheet.title = "Supply"
cellShift = 2

totalIdCount = getLastDbId(mysqlCred)
imgCounter = 0
idCounter = startExportId

mainSheet["A1"] = "ID"
mainSheet["B1"] = "Наименование"
mainSheet["C1"] = "Описание"
mainSheet["D1"] = "Цена в рублях"
mainSheet["E1"] = "Цена в злотых"
imageSheet["A1"] = "ID"
imageSheet["B1"] = "Путь к изображению"
mainImageSheet["A1"] = "ID"
mainImageSheet["B1"] = "Путь к изображению"
for id in range(totalIdCount):
    cellName = "A" + str(id + cellShift)
    mainSheet[cellName] = idCounter
    item = getItemInfo(id, mysqlCred)
    cellName = "B" + str(id + cellShift)
    mainSheet[cellName] = item['name']
    cellName = "C" + str(id + cellShift)
    mainSheet[cellName] = item['description']
    cellName = "D" + str(id + cellShift)
    mainSheet[cellName] = item['price_rub']
    cellName = "E" + str(id + cellShift)
    mainSheet[cellName] = item['price_zl']
    cellName = "A" + str(id + cellShift)
    mainImageSheet[cellName] = idCounter
    cellName = "B" + str(id + cellShift)
    mainImageSheet[cellName] = item['imgs_list'][0]
    for img in item['imgs_list']:
        cellName = "A" + str(cellShift + imgCounter)
        imageSheet[cellName] = idCounter
        cellName = "B" + str(cellShift + imgCounter)
        imageSheet[cellName] = img
        imgCounter += 1
    idCounter += 1
excelBookHandler.save("excelParsed.xlsx")



